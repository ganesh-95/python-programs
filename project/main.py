# -*- coding: utf-8 -*-
"""

"""
import os
from datetime import datetime
from googleapiclient.discovery import build
from openpyxl import load_workbook


"""
    run
"""
def run():

    make_dir('results')

    wb = load_workbook(filename = 'Book4.xlsx')
    sheet = wb[wb.get_sheet_names()[0]]
    items_count = sheet.max_row

    for rowOfCellObjects in sheet['B1':('B%s' % items_count)]:

        res = []

        for cellObj in rowOfCellObjects:

            if cellObj.value:
                print("Searching for: '%s'" % cellObj.value)

                google_search = None
                try:
                    google_search = search(cellObj.value)
                except Exception as e:
                    pass

                if google_search:
                    for query_result in google_search:
                       res.append(query_result)

                    output_file_name = ''.join(['results/output_', cellObj.coordinate, '.txt'])
                    with open(output_file_name, 'a+') as _f:
                        _f.write(''.join(res).encode('utf-8'))

                    print("...done. (file: %s)" % os.path.abspath(output_file_name))


"""
    run
"""
def search(q='', start=1):

    if not q:
        return False

    service = build("customsearch", "v1",
                    developerKey="AIzaSyBrd5AlSrZr8a_UgCS8hGRebyGu6IaXLj8")

    output = []
    try:

        res = service.cse().list(q=q,
                                 cx='016365522504479574485:an1hccvltpa',
                                 googlehost='google.com',
                                 start=start
        ).execute()

        total = int(res.get('searchInformation').get('totalResults'))

        if total > 0:
            from pprint import pprint
            for items in res.get('items'):
                output.append("%s\n%s\n%s\n\n" % (items.get('title'),
                                                  items.get('link'),
                                                  items.get('snippet')))
        else:
            pass



        return output

    except Exception as e:
        print("Error: %s" % e)
        exit(0)


def unique_filename(filename):
    file_filename = str(filename.rsplit('.', 1)[0])
    file_extension = str(filename.rsplit('.', 1)[1]).lower()
    return ''.join((file_filename, '_', datetime.now().strftime("%Y%m%d%H%M%S"), '.', file_extension))


def make_dir(dir_path):
    try:
        if not os.path.exists(dir_path):
            os.makedirs(dir_path)
    except Exception, e:
        raise e

"""
    main
"""
if __name__ == "__main__":
    run()
